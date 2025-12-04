import openpyxl

def get_data_from_excel(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    
    for r in range(2, total_rows + 1):
        row_list = []
        for c in range(1, total_cols + 1):
            cell_value = sheet.cell(row=r, column=c).value
            # FIX: Agar cell khali hai (None), to usse empty string bana do
            if cell_value is None:
                row_list.append("")
            else:
                row_list.append(cell_value)
        final_list.append(row_list)
        
    return final_list